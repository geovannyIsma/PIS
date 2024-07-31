import locale
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
from django.contrib.auth import login, logout, authenticate, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.db import IntegrityError
from django.db.models import Sum
from django.http import HttpResponseBadRequest
from django.http.response import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl.styles import Alignment, PatternFill

from Prediccion.decorators import admin_required
from Prediccion.forms import MallaCurricularForm, ExcelUploadForm, PeriodoForm, CustomUserCreationForm, \
    CustomUserChangeForm, HistoricoPeriodoForm, FeedbackForm
from Prediccion.models import MallaCurricular, Ciclo, PeriodoAcademico, Historico, CustomUser, \
    Historico_Periodo, Feedback


def index(request):
    return render(request, 'index.html')


def signin(request):
    if request.method == 'GET':
        form = AuthenticationForm()
        return render(request, 'signin.html', {'form': form})
    else:
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                return redirect('home')
            else:
                return render(request, 'signin.html', {'form': form, 'error': 'Username or password incorrect'})
        return render(request, 'signin.html', {'form': form, 'error': 'Username or password incorrect'})


def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.role = CustomUser.CONSULTOR  # Asignar el rol por defecto
                user.save()
                username = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password1')
                user = authenticate(request, username=username, password=password)
                if user is not None:
                    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                    return redirect('home')
                else:
                    return render(request, 'signup.html',
                                  {'form': form, 'error': 'Autenticación fallida. Inténtalo de nuevo.'})
            except IntegrityError:
                form.add_error(None, 'El nombre de usuario ya existe.')
        return render(request, 'signup.html', {'form': form})
    else:
        form = CustomUserCreationForm()
    return render(request, 'signup.html', {'form': form})


@login_required
def signout(request):
    logout(request)
    return redirect('index')


@admin_required
def administracion_malla(request):
    return render(request, 'administracion_malla.html')


@admin_required
def list_malla(_request):
    mallas_Curricular = list(MallaCurricular.objects.values())
    data = {'mallas_Curricular': mallas_Curricular}
    return JsonResponse(data)


@admin_required
def nueva_malla(request):
    if request.method == 'POST':
        malla_form = MallaCurricularForm(request.POST)
        if malla_form.is_valid():
            malla = malla_form.save(commit=False)

            num_cycles = int(request.POST.get('num_cycles'))
            ciclos_data = []

            for ciclo_index in range(1, num_cycles + 1):
                ciclo_data = {
                    'nombre_ciclo': f'Ciclo {ciclo_index}',
                }
                ciclos_data.append(ciclo_data)

            request.session['malla_data'] = {
                'codigo': malla.codigo,
                'nombre_malla': malla.nombre_malla,
                'tituloOtorgado': malla.tituloOtorgado
            }
            request.session['ciclos_data'] = ciclos_data

            return redirect('confirmar_malla')

    else:
        malla_form = MallaCurricularForm()

    return render(request, 'nueva_malla.html', {'malla_form': malla_form})


@admin_required
def confirmar_malla(request):
    if request.method == 'POST':
        malla_data = request.session.get('malla_data')
        ciclos_data = request.session.get('ciclos_data')

        if malla_data and ciclos_data:
            try:
                malla = MallaCurricular.objects.create(
                    codigo=malla_data['codigo'],
                    nombre_malla=malla_data['nombre_malla'],
                    tituloOtorgado=malla_data['tituloOtorgado']
                )

                for ciclo_data in ciclos_data:
                    Ciclo.objects.create(
                        nombre_ciclo=ciclo_data['nombre_ciclo'],
                        malla_curricular=malla
                    )

                request.session.pop('malla_data')
                request.session.pop('ciclos_data')

                return render(request, 'confirmar_malla.html', {'success': True})
            except Exception as e:
                return render(request, 'confirmar_malla.html', {'error': str(e)})
        else:
            return HttpResponseBadRequest("No se encontraron datos de malla o ciclos en la sesión.")
    else:
        malla_data = request.session.get('malla_data')
        ciclos_data = request.session.get('ciclos_data')

        if malla_data and ciclos_data:
            return render(request, 'confirmar_malla.html', {
                'malla_data': malla_data,
                'ciclos_data': ciclos_data
            })
        else:
            return HttpResponseBadRequest("No se encontraron datos de malla o ciclos en la sesión.")


@admin_required
def editar_malla(request, malla_id):
    malla = get_object_or_404(MallaCurricular, id=malla_id)

    if request.method == 'POST':
        malla_form = MallaCurricularForm(request.POST, instance=malla)

        if malla_form.is_valid():
            malla_form.save()

            for ciclo in malla.ciclo_set.all():
                nombre_ciclo = request.POST.get(f'nombre_ciclo_{ciclo.id}')
                ciclo.nombre_ciclo = nombre_ciclo
                ciclo.save()

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': malla_form.errors})

    else:
        malla_form = MallaCurricularForm(instance=malla)

    ciclos_data = [
        {
            'id': ciclo.id,
            'nombre_ciclo': ciclo.nombre_ciclo
        } for ciclo in malla.ciclo_set.all()
    ]

    return render(request, 'editar_malla.html', {
        'malla_form': malla_form,
        'malla': malla,
        'ciclos_data': ciclos_data,
    })


@admin_required
@require_http_methods(["DELETE"])
def eliminar_malla(request, malla_id):
    malla = get_object_or_404(MallaCurricular, id=malla_id)
    malla.delete()
    return JsonResponse({'message': 'Malla eliminada correctamente'}, status=200)


@admin_required
@csrf_exempt
def procesar_excel(request):
    if request.method == 'POST':
        file = request.FILES.get('archivo')
        if not file:
            return JsonResponse({'success': False, 'message': 'No se ha seleccionado ningún archivo.'})

        try:
            df = pd.read_excel(file, header=0, skiprows=4)
            valid_dates = df.dropna(subset=['fecha_inicio', 'fecha_fin'])
            if valid_dates.empty:
                return JsonResponse({'success': False, 'message': 'No se encontraron fechas válidas en el archivo.'})

            fecha_inicio = pd.to_datetime(valid_dates.iloc[-1]['fecha_inicio']).date()
            fecha_fin = pd.to_datetime(valid_dates.iloc[-1]['fecha_fin']).date()

            malla_id = request.POST.get('malla_id')
            if not malla_id:
                return JsonResponse({'success': False, 'message': 'Malla curricular no seleccionada.'})
            malla = MallaCurricular.objects.get(id=malla_id)

            ciclos_validos = Ciclo.objects.filter(malla_curricular=malla).values_list('nombre_ciclo', flat=True)

            historico_data = df.to_dict('records')
            historico = []
            for row in historico_data:
                if row['ciclo'] not in ciclos_validos:
                    return JsonResponse({'success': False,
                                         'message': f"El ciclo {row['ciclo']} no es válido para la malla seleccionada."})
                historico.append(
                    {
                        'ciclo': row['ciclo'],
                        'matriculados': row['matriculados'],
                        'aprobados': row['aprobados'],
                        'reprobados': row['reprobados'],
                        'abandonaron': row['abandonaron'],
                        'desertores': row['desertores']
                    }
                )

            response_data = {
                'success': True,
                'fecha_inicio': fecha_inicio.isoformat(),
                'fecha_fin': fecha_fin.isoformat(),
                'historico': historico,
                'message': 'Archivo procesado exitosamente.'
            }
            return JsonResponse(response_data)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    else:
        return JsonResponse({'success': False, 'message': 'Método no permitido.'})


@admin_required
def importar_datos_periodo_historico(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        historico_periodo_form = HistoricoPeriodoForm(request.POST)

        if form.is_valid():
            malla = form.cleaned_data['malla']
            archivo_excel = request.FILES['archivo_excel']

            fs = FileSystemStorage()
            filename = fs.save(archivo_excel.name, archivo_excel)
            file_url = fs.url(filename)

            try:
                df = pd.read_excel(fs.path(filename), header=0, skiprows=4)
                valid_dates = df.dropna(subset=['fecha_inicio', 'fecha_fin'])
                if valid_dates.empty:
                    return JsonResponse(
                        {'success': False, 'message': 'No se encontraron fechas válidas en el archivo.'})

                fecha_inicio = pd.to_datetime(valid_dates.iloc[-1]['fecha_inicio']).date()
                fecha_fin = pd.to_datetime(valid_dates.iloc[-1]['fecha_fin']).date()

                # Configurar el locale en español
                locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

                codigo_periodo = f"{fecha_inicio.strftime('%B %Y')} - {fecha_fin.strftime('%B %Y')}"

                # Restaurar el locale por defecto
                locale.setlocale(locale.LC_TIME, 'C')

                periodo, created = PeriodoAcademico.objects.update_or_create(
                    codigo_periodo=codigo_periodo,
                    defaults={'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin}
                )

                ciclos_validos = Ciclo.objects.filter(malla_curricular=malla).values_list('nombre_ciclo', flat=True)

                historicos = []
                for _, row in df.iterrows():
                    if row['ciclo'] not in ciclos_validos:
                        return JsonResponse({'success': False,
                                             'message': f"El ciclo {row['ciclo']} no es válido para la malla seleccionada."})

                    ciclo = Ciclo.objects.filter(nombre_ciclo=row['ciclo'], malla_curricular=malla).first()

                    if ciclo:
                        historico = Historico(
                            matriculados=row['matriculados'],
                            reprobados=row['reprobados'],
                            abandonaron=row['abandonaron'],
                            aprobados=row['aprobados'],
                            desertores=row['desertores'],
                            ciclo=ciclo,
                            periodo_academico=periodo
                        )
                        try:
                            historico.clean()
                            historicos.append(historico)
                        except ValidationError as e:
                            return JsonResponse(
                                {'success': False, 'message': f"Error en el ciclo {ciclo.nombre_ciclo}: {e}"})

                Historico.objects.bulk_create(historicos)

                if historico_periodo_form.is_valid():
                    historico_periodo = historico_periodo_form.save(commit=False)
                    historico_periodo.periodo_academico = periodo
                    historico_periodo.save()

                return JsonResponse({'success': True, 'message': 'Datos importados exitosamente.'})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'Error al procesar el archivo: {e}'})
            finally:
                fs.delete(filename)
        else:
            return JsonResponse({'success': False, 'message': 'Formulario no válido.'})
    else:
        form = ExcelUploadForm()
        historico_periodo_form = HistoricoPeriodoForm()

    mallas = MallaCurricular.objects.all()
    return render(request, 'import_data.html',
                  {'form': form, 'historico_periodo_form': historico_periodo_form, 'mallas': mallas})


@admin_required
def list_periodos_historicos(request):
    periodos = list(PeriodoAcademico.objects.values())
    data = {'periodos': periodos}
    return JsonResponse(data)


@admin_required
def administracion_periodo(request):
    return render(request, 'administracion_registros_periodos.html')


@login_required
def editar_datos_periodo_historico(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id)
    historico_periodo = get_object_or_404(Historico_Periodo, periodo_academico=periodo)

    if request.method == 'POST':
        periodo_form = PeriodoForm(request.POST, instance=periodo)
        historico_periodo_form = HistoricoPeriodoForm(request.POST, instance=historico_periodo)

        if periodo_form.is_valid() and historico_periodo_form.is_valid():
            periodo_form.save()
            historico_periodo_form.save()

            historicos = []
            for historico in periodo.historico_set.all():
                historico.matriculados = request.POST.get(f'matriculados_{historico.id}')
                historico.reprobados = request.POST.get(f'reprobados_{historico.id}')
                historico.abandonaron = request.POST.get(f'abandonaron_{historico.id}')
                historico.aprobados = request.POST.get(f'aprobados_{historico.id}')
                historico.desertores = request.POST.get(f'desertores_{historico.id}')

                try:
                    historico.clean()
                    historicos.append(historico)
                except ValidationError as e:
                    return JsonResponse(
                        {'success': False, 'message': f"Error en el ciclo {historico.ciclo.nombre_ciclo}: {e}"})

            Historico.objects.bulk_update(historicos,
                                          ['matriculados', 'reprobados', 'abandonaron', 'aprobados', 'desertores'])

            return JsonResponse({'success': True, 'message': 'Datos del periodo actualizados exitosamente.'})
        else:
            return JsonResponse({'success': False, 'message': 'Formulario no válido.'})
    else:
        periodo.fecha_inicio = periodo.fecha_inicio.strftime('%Y-%m-%d')
        periodo.fecha_fin = periodo.fecha_fin.strftime('%Y-%m-%d')
        periodo_form = PeriodoForm(instance=periodo)
        historico_periodo_form = HistoricoPeriodoForm(instance=historico_periodo)

    historicos = periodo.historico_set.all()
    context = {
        'periodo': periodo,
        'periodo_form': periodo_form,
        'historico_periodo_form': historico_periodo_form,
        'historicos': historicos
    }
    return render(request, 'editar_periodo.html', context)


@admin_required
def eliminar_datos_periodo_historico(request, periodo_id):
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id)
    periodo.delete()
    return JsonResponse({'message': 'Periodo eliminado correctamente'}, status=200)


def modelo_matematico(request):
    return render(request, 'modelo_matematico.html')


User = get_user_model()


@admin_required
def administracion_usuarios(request):
    return render(request, 'administracion_usuarios.html')


@admin_required
def list_usuarios(request):
    usuarios = list(User.objects.filter(is_superuser=False, is_staff=False).values())
    data = {'usuarios': usuarios}
    return JsonResponse(data)


@admin_required
def nuevo_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('administracion_usuarios')
    else:
        form = CustomUserCreationForm()
    return render(request, 'nuevo_usuario.html', {'form': form})


@admin_required
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id, is_superuser=False, is_staff=False)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('administracion_usuario')
    else:
        form = CustomUserChangeForm(instance=usuario)
    return render(request, 'edit_user.html', {'form': form, 'usuario': usuario})


@admin_required
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(User, id=usuario_id, is_superuser=False, is_staff=False)
    usuario.delete()
    return JsonResponse({'message': 'Usuario eliminado correctamente'}, status=200)


def dashboard_view(request):
    periodo_seleccionado_id = request.GET.get('periodo', None)

    periodos_academicos = PeriodoAcademico.objects.all()

    if periodo_seleccionado_id:
        historicos = Historico.objects.filter(periodo_academico_id=periodo_seleccionado_id).order_by(
            'periodo_academico__fecha_inicio')
        historico_periodo = Historico_Periodo.objects.filter(periodo_academico_id=periodo_seleccionado_id)
    else:
        historicos = Historico.objects.order_by('periodo_academico__fecha_inicio')
        historico_periodo = Historico_Periodo.objects.all()

    periodos = [h.periodo_academico.codigo_periodo for h in historicos]
    matriculados = [h.matriculados for h in historicos]
    desertores = [h.desertores for h in historicos]

    # Datos para el Historico_Periodo
    periodo_matriculados = sum([h.matriculados for h in historico_periodo])
    periodo_reprobados = sum([h.reprobados for h in historico_periodo])
    periodo_abandonaron = sum([h.abandonaron for h in historico_periodo])
    periodo_aprobados = sum([h.aprobados for h in historico_periodo])
    periodo_desertores = sum([h.desertores for h in historico_periodo])

    total_matriculados = historicos.aggregate(Sum('matriculados'))['matriculados__sum'] or 0
    total_reprobados = historicos.aggregate(Sum('reprobados'))['reprobados__sum'] or 0
    total_abandonaron = historicos.aggregate(Sum('abandonaron'))['abandonaron__sum'] or 0
    total_aprobados = historicos.aggregate(Sum('aprobados'))['aprobados__sum'] or 0
    total_desertores = historicos.aggregate(Sum('desertores'))['desertores__sum'] or 0

    context = {
        'periodos_academicos': periodos_academicos,
        'periodo_seleccionado_id': periodo_seleccionado_id,
        'periodos': periodos,
        'matriculados': matriculados,
        'desertores': desertores,
        'periodo_matriculados': periodo_matriculados,
        'periodo_reprobados': periodo_reprobados,
        'periodo_abandonaron': periodo_abandonaron,
        'periodo_aprobados': periodo_aprobados,
        'periodo_desertores': periodo_desertores,
    }

    return render(request, 'home.html', context)


def about(request):
    return render(request, 'about.html')


"""
def mcmc_cadena(lista):
    difer = np.diff(lista)
    media = np.mean(difer)
    desviacion = np.std(difer)
    sim = 5000
    valores_futuros = []
    for _ in range(sim):
        nueva_diferencia = np.random.normal(media, desviacion)
        nuevo_valor = lista[-1] + nueva_diferencia
        valores_futuros.append(nuevo_valor)
    valor_futuro_estimado = np.mean(valores_futuros)
    return valor_futuro_estimado


def format_periodo(fecha_inicio, fecha_fin):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    inicio = fecha_inicio.strftime('%B %Y')
    fin = fecha_fin.strftime('%B %Y')
    return f"{inicio} - {fin}"


def generar_periodos_futuros(ultimo_periodo_inicio, ultimo_periodo_fin, cantidad_futuros):
    periodos_futuros = []
    fecha_inicio = ultimo_periodo_inicio
    fecha_fin = ultimo_periodo_fin

    for _ in range(cantidad_futuros):
        fecha_inicio += relativedelta(months=6)
        fecha_fin += relativedelta(months=6)
        periodos_futuros.append(format_periodo(fecha_inicio, fecha_fin))

    return periodos_futuros


def indiciar(prediccion, historicos):
    resul = historicos.copy()
    for i in range(10, len(prediccion), 10):
        resul.append(round(prediccion[i]))
    return resul


def prediccion_view(request):
    historico = Historico.objects.all()
    matriculados = list(historico.values_list('matriculados', flat=True))
    aprobados = list(historico.values_list('aprobados', flat=True))
    reprobados = list(historico.values_list('reprobados', flat=True))
    desertores = list(historico.values_list('desertores', flat=True))
    abandonos = list(historico.values_list('abandonaron', flat=True))

    lmbda = mcmc_cadena(desertores) / mcmc_cadena(matriculados)
    beta = mcmc_cadena(aprobados) / mcmc_cadena(matriculados)
    gamma = mcmc_cadena(reprobados) / mcmc_cadena(matriculados)
    alpha = mcmc_cadena(abandonos) / mcmc_cadena(reprobados)

    def dMdt(t, M, R, D, A, AB):
        return -lmbda * M * AB

    def dAdt(t, M, R, D, A, AB):
        return beta * M - alpha * A

    def dRdt(t, M, R, D, A, AB):
        return gamma * M - alpha * R - lmbda * R

    def dDdt(t, M, R, D, A, AB):
        return lmbda * M + lmbda * R

    def dABdt(t, M, R, D, A, AB):
        return alpha * A + alpha * R

    def rungeKutta4toOrden(t, mi, ai, ri, di, abi, h):
        solMatriculados = [mi]
        solaprobados = [ai]
        solReprobados = [ri]
        solDesertores = [di]
        solAbandono = [abi]
        tiempo = [t]

        for i in range(200):
            M1 = dMdt(t, mi, ri, di, ai, abi)
            A1 = dAdt(t, mi, ri, di, ai, abi)
            R1 = dRdt(t, mi, ri, di, ai, abi)
            D1 = dDdt(t, mi, ri, di, ai, abi)
            AB1 = dABdt(t, mi, ri, di, ai, abi)

            M2 = dMdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2)
            A2 = dAdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2)
            R2 = dRdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2)
            D2 = dDdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2)
            AB2 = dABdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2,
                        abi + AB1 * h / 2)

            M3 = dMdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2)
            A3 = dAdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2)
            R3 = dRdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2)
            D3 = dDdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2)
            AB3 = dABdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2,
                        abi + AB2 * h / 2)

            M4 = dMdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h)
            A4 = dAdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h)
            R4 = dRdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h)
            D4 = dDdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h)
            AB4 = dABdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h)

            mi = mi + (M1 + 2 * M2 + 2 * M3 + M4) * h / 6
            ai = ai + (A1 + 2 * A2 + 2 * A3 + A4) * h / 6
            ri = ri + (R1 + 2 * R2 + 2 * R3 + R4) * h / 6
            di = di + (D1 + 2 * D2 + 2 * D3 + D4) * h / 6
            abi = abi + (AB1 + 2 * AB2 + 2 * AB3 + AB4) * h / 6
            t = t + h

            solMatriculados.append(mi)
            solaprobados.append(ai)
            solReprobados.append(ri)
            solDesertores.append(di)
            solAbandono.append(abi)
            tiempo.append(t)

        return solMatriculados, solaprobados, solReprobados, solDesertores, solAbandono, tiempo

    solMatriculados, solaprobados, solReprobados, solDesertores, solAbandono, tiempo = rungeKutta4toOrden(
        0, matriculados[-1], aprobados[-1], reprobados[-1], desertores[-1], abandonos[-1], 0.1)

    historicos = Historico.objects.select_related('periodo_academico').order_by('periodo_academico__fecha_inicio')
    periodos_historicos = [format_periodo(h.periodo_academico.fecha_inicio, h.periodo_academico.fecha_fin) for h in
                           historicos]
    ultimo_periodo = historicos.last().periodo_academico
    periodos_futuros = generar_periodos_futuros(ultimo_periodo.fecha_inicio, ultimo_periodo.fecha_fin,
                                                len(tiempo) - len(periodos_historicos))

    # Utilizar la función indiciar para integrar los datos predichos
    solMatriculados = indiciar(solMatriculados, matriculados)
    solaprobados = indiciar(solaprobados, aprobados)
    solReprobados = indiciar(solReprobados, reprobados)
    solDesertores = indiciar(solDesertores, desertores)
    solAbandono = indiciar(solAbandono, abandonos)

    data = {
        'tiempo': tiempo,
        'matriculados': solMatriculados,
        'aprobados': solaprobados,
        'reprobados': solReprobados,
        'desertores': solDesertores,
        'abandono': solAbandono,
        'periodos': periodos_historicos + periodos_futuros
    }
    return render(request, 'prediccion.html', {'data': data})
"""


def simulacion_view(request):
    return render(request, 'simulacion.html')


def registros_almacenados(request):
    periodos = PeriodoAcademico.objects.all()

    if request.method == 'POST':
        periodos_seleccionados = request.POST.getlist('periodos')

        if periodos_seleccionados:
            output = BytesIO()
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Registros Almacenados"

            row_num = 1
            fill_colors = ['FFC7CE', 'C6EFCE', 'FFEB9C', 'D9EAD3', 'FCE4D6']  # Lista de colores de fondo
            color_index = 0

            for periodo_id in periodos_seleccionados:
                fill_color = PatternFill(start_color=fill_colors[color_index % len(fill_colors)],
                                         end_color=fill_colors[color_index % len(fill_colors)], fill_type="solid")
                color_index += 1

                periodo = get_object_or_404(PeriodoAcademico, id=periodo_id)
                historico = Historico.objects.filter(periodo_academico=periodo)
                historico_periodo = Historico_Periodo.objects.filter(periodo_academico=periodo)

                # Escribir Periodo
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                cell = sheet.cell(row=row_num, column=1)
                cell.value = f"Periodo: {periodo.codigo_periodo}"
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill_color
                row_num += 1
                sheet.append(['Código Período', 'Fecha Inicio', 'Fecha Fin'])
                for col in range(1, 4):
                    sheet.cell(row=row_num, column=col).fill = fill_color
                sheet.append([
                    periodo.codigo_periodo,
                    periodo.fecha_inicio.strftime('%Y-%m-%d'),
                    periodo.fecha_fin.strftime('%Y-%m-%d')
                ])
                for col in range(1, 4):
                    sheet.cell(row=row_num, column=col).fill = fill_color
                row_num += 2

                # Escribir Historico
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                cell = sheet.cell(row=row_num, column=1)
                cell.value = "Historico"
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill_color
                row_num += 1
                sheet.append(['Ciclo', 'Matriculados', 'Aprobados', 'Reprobados', 'Abandonaron', 'Desertores'])
                for col in range(1, 7):
                    sheet.cell(row=row_num, column=col).fill = fill_color
                for record in historico:
                    row_num += 1
                    sheet.append([
                        record.ciclo.nombre_ciclo,
                        record.matriculados,
                        record.aprobados,
                        record.reprobados,
                        record.abandonaron,
                        record.desertores
                    ])
                    for col in range(1, 7):
                        sheet.cell(row=row_num, column=col).fill = fill_color
                row_num += 2

                # Escribir Historico_Periodo
                sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
                cell = sheet.cell(row=row_num, column=1)
                cell.value = "Historico Periodo"
                cell.alignment = Alignment(horizontal='center')
                cell.fill = fill_color
                row_num += 1
                sheet.append(['Periodo', 'Matriculados', 'Aprobados', 'Reprobados', 'Abandonaron', 'Desertores'])
                for col in range(1, 7):
                    sheet.cell(row=row_num, column=col).fill = fill_color
                for record in historico_periodo:
                    row_num += 1
                    sheet.append([
                        record.periodo_academico.codigo_periodo,
                        record.matriculados,
                        record.aprobados,
                        record.reprobados,
                        record.abandonaron,
                        record.desertores
                    ])
                    for col in range(1, 7):
                        sheet.cell(row=row_num, column=col).fill = fill_color
                row_num += 3

            workbook.save(output)
            output.seek(0)

            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=registros_periodos.xlsx'
            return response

    context = {
        'periodos': periodos
    }
    return render(request, 'registros_almacenados.html', context)


@login_required
def enviar_feedback(request):
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.usuario = request.user
            feedback.save()
            return redirect(f'/home/?feedback=success')
    else:
        form = FeedbackForm()
    return render(request, 'enviar_feedback.html', {'form': form})


@admin_required
def leer_feedback(request):
    feedbacks = Feedback.objects.all().order_by('-fecha_creacion')
    return render(request, 'leer_feedback.html', {'feedbacks': feedbacks})


def mcmc_cadena(lista, num_sim=5000):
    if len(lista) < 2:
        # Añadir variabilidad para listas muy pequeñas o constantes
        base_valor = lista[-1] if lista else 0
        variabilidad = base_valor * 0.05  # 5% de variabilidad
        valores_futuros = [base_valor + np.random.uniform(-variabilidad, variabilidad) for _ in range(num_sim)]
    else:
        difer = np.diff(lista)
        media = np.mean(difer)
        desviacion = np.std(difer)

        if desviacion == 0:  # Caso de datos constantes
            variabilidad = media * 0.05  # 5% de variabilidad
            valores_futuros = [lista[-1] + media + np.random.uniform(-variabilidad, variabilidad) for _ in
                               range(num_sim)]
        else:
            valores_futuros = []
            for _ in range(num_sim):
                nueva_diferencia = np.random.normal(media, desviacion)
                nuevo_valor = lista[-1] + nueva_diferencia
                valores_futuros.append(nuevo_valor)
    return valores_futuros


def predicciones_view(request):
    # Configurar el locale en español
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    # Obtener lista de ciclos disponibles
    ciclos = Ciclo.objects.all()

    # Obtener ciclo seleccionado
    ciclo_id = request.GET.get('ciclo', ciclos.first().id)
    ciclo_seleccionado = get_object_or_404(Ciclo, id=ciclo_id)

    # Filtrar datos históricos por ciclo seleccionado
    historicos = Historico.objects.filter(ciclo=ciclo_seleccionado).order_by('id')

    if historicos.exists():
        periodos_ciclo = [h.periodo_academico.codigo_periodo for h in historicos]
        matriculados_ciclo = [h.matriculados for h in historicos]
        reprobados_ciclo = [h.reprobados for h in historicos]
        abandonaron_ciclo = [h.abandonaron for h in historicos]
        aprobados_ciclo = [h.aprobados for h in historicos]
        desertores_ciclo = [h.desertores for h in historicos]

        # Realizar predicciones por ciclo
        prediccion_matriculados_ciclo = mcmc_cadena(matriculados_ciclo)
        prediccion_reprobados_ciclo = mcmc_cadena(reprobados_ciclo)
        prediccion_abandonaron_ciclo = mcmc_cadena(abandonaron_ciclo)
        prediccion_aprobados_ciclo = mcmc_cadena(aprobados_ciclo)
        prediccion_desertores_ciclo = mcmc_cadena(desertores_ciclo)

        # Generar nuevos periodos para ciclo
        ultimos_periodo_ciclo = historicos.last().periodo_academico
        fecha_inicio = ultimos_periodo_ciclo.fecha_fin
        nuevos_periodos_ciclo = [
            f"{(fecha_inicio + pd.DateOffset(months=i * 6)).strftime('%B %Y')} - {(fecha_inicio + pd.DateOffset(months=(i + 1) * 6 - 1)).strftime('%B %Y')}"
            for i in range(12)
        ]
    else:
        periodos_ciclo = []
        matriculados_ciclo = []
        reprobados_ciclo = []
        abandonaron_ciclo = []
        aprobados_ciclo = []
        desertores_ciclo = []
        nuevos_periodos_ciclo = []
        prediccion_matriculados_ciclo = []
        prediccion_reprobados_ciclo = []
        prediccion_abandonaron_ciclo = []
        prediccion_aprobados_ciclo = []
        prediccion_desertores_ciclo = []

    # Obtener datos históricos por período académico
    historicos_periodo = Historico_Periodo.objects.all().order_by('periodo_academico__fecha_inicio')

    if historicos_periodo.exists():
        periodos_periodo = [h.periodo_academico.codigo_periodo for h in historicos_periodo]
        matriculados_periodo = [h.matriculados for h in historicos_periodo]
        reprobados_periodo = [h.reprobados for h in historicos_periodo]
        abandonaron_periodo = [h.abandonaron for h in historicos_periodo]
        aprobados_periodo = [h.aprobados for h in historicos_periodo]
        desertores_periodo = [h.desertores for h in historicos_periodo]

        # Realizar predicciones por período académico
        prediccion_matriculados_periodo = mcmc_cadena(matriculados_periodo)
        prediccion_reprobados_periodo = mcmc_cadena(reprobados_periodo)
        prediccion_abandonaron_periodo = mcmc_cadena(abandonaron_periodo)
        prediccion_aprobados_periodo = mcmc_cadena(aprobados_periodo)
        prediccion_desertores_periodo = mcmc_cadena(desertores_periodo)

        # Generar nuevos periodos para período académico
        ultimos_periodo_periodo = historicos_periodo.last().periodo_academico
        fecha_inicio = ultimos_periodo_periodo.fecha_fin
        nuevos_periodos_periodo = [
            f"{(fecha_inicio + pd.DateOffset(months=i * 6)).strftime('%B %Y')} - {(fecha_inicio + pd.DateOffset(months=(i + 1) * 6 - 1)).strftime('%B %Y')}"
            for i in range(12)
        ]
    else:
        periodos_periodo = []
        matriculados_periodo = []
        reprobados_periodo = []
        abandonaron_periodo = []
        aprobados_periodo = []
        desertores_periodo = []
        nuevos_periodos_periodo = []
        prediccion_matriculados_periodo = []
        prediccion_reprobados_periodo = []
        prediccion_abandonaron_periodo = []
        prediccion_aprobados_periodo = []
        prediccion_desertores_periodo = []

    context = {
        'ciclos': ciclos,
        'ciclo_seleccionado': ciclo_seleccionado,
        'periodos_ciclo': periodos_ciclo,
        'matriculados_ciclo': matriculados_ciclo,
        'reprobados_ciclo': reprobados_ciclo,
        'abandonaron_ciclo': abandonaron_ciclo,
        'aprobados_ciclo': aprobados_ciclo,
        'desertores_ciclo': desertores_ciclo,
        'nuevos_periodos_ciclo': nuevos_periodos_ciclo,
        'prediccion_matriculados_ciclo': prediccion_matriculados_ciclo,
        'prediccion_reprobados_ciclo': prediccion_reprobados_ciclo,
        'prediccion_abandonaron_ciclo': prediccion_abandonaron_ciclo,
        'prediccion_aprobados_ciclo': prediccion_aprobados_ciclo,
        'prediccion_desertores_ciclo': prediccion_desertores_ciclo,
        'periodos_periodo': periodos_periodo,
        'matriculados_periodo': matriculados_periodo,
        'reprobados_periodo': reprobados_periodo,
        'abandonaron_periodo': abandonaron_periodo,
        'aprobados_periodo': aprobados_periodo,
        'desertores_periodo': desertores_periodo,
        'nuevos_periodos_periodo': nuevos_periodos_periodo,
        'prediccion_matriculados_periodo': prediccion_matriculados_periodo,
        'prediccion_reprobados_periodo': prediccion_reprobados_periodo,
        'prediccion_abandonaron_periodo': prediccion_abandonaron_periodo,
        'prediccion_aprobados_periodo': prediccion_aprobados_periodo,
        'prediccion_desertores_periodo': prediccion_desertores_periodo,
    }

    # Restaurar el locale por defecto
    locale.setlocale(locale.LC_TIME, 'C')

    return render(request, 'prediccion.html', context)


# Prediccion usando modelo matemático
def dMdt(t, M, R, D, A, AB, lmbda):
    return -lmbda * M * AB


def dAdt(t, M, R, D, A, AB, beta, alpha):
    return beta * M - alpha * A


def dRdt(t, M, R, D, A, AB, gamma, alpha, lmbda):
    return gamma * M - alpha * R - lmbda * R


def dDdt(t, M, R, D, A, AB, lmbda):
    return lmbda * M + lmbda * R


def dABdt(t, M, R, D, A, AB, alpha):
    return alpha * A + alpha * R


# Implementación del método de Runge-Kutta de 4º orden
def rungeKutta4toOrden(t, mi, ai, ri, di, abi, h, lmbda, beta, gamma, alpha):
    solMatriculados = [mi]
    solaprobados = [ai]
    solReprobados = [ri]
    solDesertores = [di]
    solAbandono = [abi]
    tiempo = [t]

    for i in range(200):
        M1 = dMdt(t, mi, ri, di, ai, abi, lmbda)
        A1 = dAdt(t, mi, ri, di, ai, abi, beta, alpha)
        R1 = dRdt(t, mi, ri, di, ai, abi, gamma, alpha, lmbda)
        D1 = dDdt(t, mi, ri, di, ai, abi, lmbda)
        AB1 = dABdt(t, mi, ri, di, ai, abi, alpha)

        M2 = dMdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2,
                  lmbda)
        A2 = dAdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2,
                  beta, alpha)
        R2 = dRdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2,
                  gamma, alpha, lmbda)
        D2 = dDdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2,
                  lmbda)
        AB2 = dABdt(t + h / 2, mi + M1 * h / 2, ri + R1 * h / 2, di + D1 * h / 2, ai + A1 * h / 2, abi + AB1 * h / 2,
                    alpha)

        M3 = dMdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2,
                  lmbda)
        A3 = dAdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2,
                  beta, alpha)
        R3 = dRdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2,
                  gamma, alpha, lmbda)
        D3 = dDdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2,
                  lmbda)
        AB3 = dABdt(t + h / 2, mi + M2 * h / 2, ri + R2 * h / 2, di + D2 * h / 2, ai + A2 * h / 2, abi + AB2 * h / 2,
                    alpha)

        M4 = dMdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h, lmbda)
        A4 = dAdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h, beta, alpha)
        R4 = dRdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h, gamma, alpha, lmbda)
        D4 = dDdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h, lmbda)
        AB4 = dABdt(t + h, mi + M3 * h, ri + R3 * h, di + D3 * h, ai + A3 * h, abi + AB3 * h, alpha)

        mi = mi + (M1 + 2 * M2 + 2 * M3 + M4) * h / 6
        ai = ai + (A1 + 2 * A2 + 2 * A3 + A4) * h / 6
        ri = ri + (R1 + 2 * R2 + 2 * R3 + R4) * h / 6
        di = di + (D1 + 2 * D2 + 2 * D3 + D4) * h / 6
        abi = abi + (AB1 + 2 * AB2 + 2 * AB3 + AB4) * h / 6
        t = t + h

        solMatriculados.append(mi)
        solaprobados.append(ai)
        solReprobados.append(ri)
        solDesertores.append(di)
        solAbandono.append(abi)
        tiempo.append(t)

    return solMatriculados, solaprobados, solReprobados, solDesertores, solAbandono, tiempo


def mcmc(lista, num_sim=5000):
    if len(lista) < 2:
        base_valor = lista[-1] if lista else 0
        variabilidad = base_valor * 0.05  # 5% de variabilidad
        valores_futuros = [base_valor + np.random.uniform(-variabilidad, variabilidad) for _ in range(num_sim)]
    else:
        difer = np.diff(lista)
        media = np.mean(difer)
        desviacion = np.std(difer)

        if desviacion == 0:  # Caso de datos constantes
            variabilidad = media * 0.05  # 5% de variabilidad
            valores_futuros = [lista[-1] + media + np.random.uniform(-variabilidad, variabilidad) for _ in
                               range(num_sim)]
        else:
            valores_futuros = []
            for _ in range(num_sim):
                nueva_diferencia = np.random.normal(media, desviacion)
                nuevo_valor = lista[-1] + nueva_diferencia
                valores_futuros.append(nuevo_valor)
    return valores_futuros


# Definición de la función de predicción usando el modelo matemático
def prediccion_modelo_matematico(matriculados, aprobados, reprobados, desertores, abandonos, h=0.1, num_periodos=12):
    # Calcular la media de los valores simulados por MCMC para cada lista
    lmbda = np.mean(mcmc(desertores)) / np.mean(mcmc(matriculados))
    beta = np.mean(mcmc(aprobados)) / np.mean(mcmc(matriculados))
    gamma = np.mean(mcmc(reprobados)) / np.mean(mcmc(matriculados))
    alpha = np.mean(mcmc(abandonos)) / np.mean(mcmc(reprobados))

    # Realiza la predicción usando el método de Runge-Kutta de 4º orden
    solMatriculados, solaprobados, solReprobados, solDesertores, solAbandono, tiempo = rungeKutta4toOrden(
        0, matriculados[-1], aprobados[-1], reprobados[-1], desertores[-1], abandonos[-1], h, lmbda, beta, gamma, alpha
    )

    # Truncar los resultados a los primeros `num_periodos`
    solMatriculados = solMatriculados[:num_periodos]
    solaprobados = solaprobados[:num_periodos]
    solReprobados = solReprobados[:num_periodos]
    solDesertores = solDesertores[:num_periodos]
    solAbandono = solAbandono[:num_periodos]
    tiempo = tiempo[:num_periodos]

    # Regresa los valores predichos para cada variable
    return solMatriculados, solaprobados, solReprobados, solDesertores, solAbandono


def prediccion_matematica_view(request):
    # Configurar el locale en español
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    # Obtener lista de ciclos disponibles
    ciclos = Ciclo.objects.all()

    # Obtener ciclo seleccionado
    ciclo_id = request.GET.get('ciclo', ciclos.first().id)
    ciclo_seleccionado = get_object_or_404(Ciclo, id=ciclo_id)

    # Filtrar datos históricos por ciclo seleccionado
    historicos = Historico.objects.filter(ciclo=ciclo_seleccionado).order_by('id')

    if historicos.exists():
        periodos_ciclo = [h.periodo_academico.codigo_periodo for h in historicos]
        matriculados_ciclo = [h.matriculados for h in historicos]
        reprobados_ciclo = [h.reprobados for h in historicos]
        abandonaron_ciclo = [h.abandonaron for h in historicos]
        aprobados_ciclo = [h.aprobados for h in historicos]
        desertores_ciclo = [h.desertores for h in historicos]

        # Realizar predicciones por ciclo usando el modelo matemático
        (prediccion_matriculados_ciclo,
         prediccion_aprobados_ciclo,
         prediccion_reprobados_ciclo,
         prediccion_desertores_ciclo,
         prediccion_abandonaron_ciclo) = prediccion_modelo_matematico(matriculados_ciclo, aprobados_ciclo,
                                                                      reprobados_ciclo, desertores_ciclo,
                                                                      abandonaron_ciclo)

        # Generar nuevos periodos para ciclo
        ultimos_periodo_ciclo = historicos.last().periodo_academico
        fecha_inicio = ultimos_periodo_ciclo.fecha_fin
        nuevos_periodos_ciclo = [
            f"{(fecha_inicio + pd.DateOffset(months=i * 6)).strftime('%B %Y')} - {(fecha_inicio + pd.DateOffset(months=(i + 1) * 6 - 1)).strftime('%B %Y')}"
            for i in range(12)
        ]
    else:
        periodos_ciclo = []
        matriculados_ciclo = []
        reprobados_ciclo = []
        abandonaron_ciclo = []
        aprobados_ciclo = []
        desertores_ciclo = []
        nuevos_periodos_ciclo = []
        prediccion_matriculados_ciclo = []
        prediccion_reprobados_ciclo = []
        prediccion_abandonaron_ciclo = []
        prediccion_aprobados_ciclo = []
        prediccion_desertores_ciclo = []

    # Obtener datos históricos por período académico
    historicos_periodo = Historico_Periodo.objects.all().order_by('periodo_academico__fecha_inicio')

    if historicos_periodo.exists():
        periodos_periodo = [h.periodo_academico.codigo_periodo for h in historicos_periodo]
        matriculados_periodo = [h.matriculados for h in historicos_periodo]
        reprobados_periodo = [h.reprobados for h in historicos_periodo]
        abandonaron_periodo = [h.abandonaron for h in historicos_periodo]
        aprobados_periodo = [h.aprobados for h in historicos_periodo]
        desertores_periodo = [h.desertores for h in historicos_periodo]

        # Realizar predicciones por período académico usando el modelo matemático
        (prediccion_matriculados_periodo,
         prediccion_aprobados_periodo,
         prediccion_reprobados_periodo,
         prediccion_desertores_periodo,
         prediccion_abandonaron_periodo) = prediccion_modelo_matematico(matriculados_periodo, aprobados_periodo,
                                                                        reprobados_periodo, desertores_periodo,
                                                                        abandonaron_periodo)

        # Generar nuevos periodos para período académico
        ultimos_periodo_periodo = historicos_periodo.last().periodo_academico
        fecha_inicio = ultimos_periodo_periodo.fecha_fin
        nuevos_periodos_periodo = [
            f"{(fecha_inicio + pd.DateOffset(months=i * 6)).strftime('%B %Y')} - {(fecha_inicio + pd.DateOffset(months=(i + 1) * 6 - 1)).strftime('%B %Y')}"
            for i in range(12)
        ]
    else:
        periodos_periodo = []
        matriculados_periodo = []
        reprobados_periodo = []
        abandonaron_periodo = []
        aprobados_periodo = []
        desertores_periodo = []
        nuevos_periodos_periodo = []
        prediccion_matriculados_periodo = []
        prediccion_reprobados_periodo = []
        prediccion_abandonaron_periodo = []
        prediccion_aprobados_periodo = []
        prediccion_desertores_periodo = []

    context = {
        'ciclos': ciclos,
        'ciclo_seleccionado': ciclo_seleccionado,
        'periodos_ciclo': periodos_ciclo,
        'matriculados_ciclo': matriculados_ciclo,
        'reprobados_ciclo': reprobados_ciclo,
        'abandonaron_ciclo': abandonaron_ciclo,
        'aprobados_ciclo': aprobados_ciclo,
        'desertores_ciclo': desertores_ciclo,
        'nuevos_periodos_ciclo': nuevos_periodos_ciclo,
        'prediccion_matriculados_ciclo': prediccion_matriculados_ciclo,
        'prediccion_reprobados_ciclo': prediccion_reprobados_ciclo,
        'prediccion_abandonaron_ciclo': prediccion_abandonaron_ciclo,
        'prediccion_aprobados_ciclo': prediccion_aprobados_ciclo,
        'prediccion_desertores_ciclo': prediccion_desertores_ciclo,
        'periodos_periodo': periodos_periodo,
        'matriculados_periodo': matriculados_periodo,
        'reprobados_periodo': reprobados_periodo,
        'abandonaron_periodo': abandonaron_periodo,
        'aprobados_periodo': aprobados_periodo,
        'desertores_periodo': desertores_periodo,
        'nuevos_periodos_periodo': nuevos_periodos_periodo,
        'prediccion_matriculados_periodo': prediccion_matriculados_periodo,
        'prediccion_reprobados_periodo': prediccion_reprobados_periodo,
        'prediccion_abandonaron_periodo': prediccion_abandonaron_periodo,
        'prediccion_aprobados_periodo': prediccion_aprobados_periodo,
        'prediccion_desertores_periodo': prediccion_desertores_periodo,
    }

    locale.setlocale(locale.LC_TIME, 'C')

    return render(request, 'prediccion_matematica.html', context)
